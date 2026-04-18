# ============================================================
# qa_tester.py — 课件后测试脚本
# 每次运行对一个 Unit 文件夹执行两轮 QA 测试：
#   Test 1: 输入 unit_outline.json → 大纲质量评估
#   Test 2: 输入 unit_outline.json + 随机3份 lesson JSON → 课件内容评估
# ============================================================

import json
import random
import re
import sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from config import AI_MODEL, client

# ── Prompt 占位符（等待填入）─────────────────────────────────

from content_processor import _load_section

OUTLINE_QA_PROMPT = _load_section("common_qa_prompts.md", "OUTLINE_QA_PROMPT")
LESSON_QA_PROMPT = _load_section("common_qa_prompts.md", "LESSON_QA_PROMPT")


# ── 程序化校验层（在 AI QA 之前先做硬性检查）─────────────────

REQUIRED_MODULES = {"title", "warm_up", "conversation_builder", "practice", "scenario", "wrap_up"}
ALLOWED_MODULES  = REQUIRED_MODULES | {"useful_language_1", "useful_language_2"}
BANNED_MODULES   = {"grammar_focus", "speaking_chain", "quick_check"}


def _normalize(text: str) -> str:
    """小写 + 去标点，用于模糊匹配"""
    return re.sub(r"[^a-z0-9\s]", "", text.lower()).strip()


def programmatic_check_lesson(lesson_data: list[dict], outline_lesson: dict) -> list[str]:
    """
    对单课 JSON 做硬性校验，返回问题列表（空 = 全部通过）。
    校验项：
      1. 模块完整性 & 违禁模块
      2. 词汇忠实度（useful_language vs outline.vocabulary）
      3. 句型忠实度（conversation_builder vs outline.functional_language）
    """
    issues = []
    slide_types = [s.get("type", "") for s in lesson_data]

    # ── 1. 模块完整性 ──
    present = set(slide_types)
    # useful_language 可能拆成 _1 和 _2，合并判断
    has_useful_lang = any(t.startswith("useful_language") for t in present)
    missing = REQUIRED_MODULES - present
    if missing:
        issues.append(f"[结构] 缺少必需模块: {', '.join(sorted(missing))}")
    if not has_useful_lang:
        issues.append("[结构] 缺少 useful_language 模块（_1 或 _2 均未找到）")

    banned_found = BANNED_MODULES & {t.lower() for t in present}
    if banned_found:
        issues.append(f"[结构] 存在违禁模块: {', '.join(sorted(banned_found))}")

    # ── 2. 词汇忠实度 ──
    outline_vocab = {_normalize(v) for v in outline_lesson.get("vocabulary", [])}
    lesson_words = set()
    for slide in lesson_data:
        if slide.get("type", "").startswith("useful_language"):
            for w in slide.get("words", []):
                lesson_words.add(_normalize(w.get("word", "")))

    if outline_vocab:
        missing_vocab = outline_vocab - lesson_words
        extra_vocab = lesson_words - outline_vocab
        if missing_vocab:
            issues.append(f"[词汇] 大纲词汇未出现在课件中: {', '.join(sorted(missing_vocab))}")
        if extra_vocab:
            issues.append(f"[词汇] 课件引入了大纲外词汇: {', '.join(sorted(extra_vocab))}")

    # ── 3. 句型忠实度 ──
    outline_fl = {_normalize(f) for f in outline_lesson.get("functional_language", [])}
    lesson_linkers = set()
    for slide in lesson_data:
        if slide.get("type") == "conversation_builder":
            for lk in slide.get("linkers", []):
                lesson_linkers.add(_normalize(lk.get("word", "")))

    if outline_fl:
        missing_fl = outline_fl - lesson_linkers
        if missing_fl:
            # 做模糊包含检查：如果 linker 文本包含了 outline 句型，也算覆盖
            still_missing = set()
            all_linker_text = " ".join(lesson_linkers)
            for fl in missing_fl:
                if fl not in all_linker_text:
                    still_missing.add(fl)
            if still_missing:
                issues.append(f"[句型] 大纲句型未在 conversation_builder 中覆盖: {', '.join(sorted(still_missing))}")

    return issues


def programmatic_check_outline(outline: dict) -> list[str]:
    """
    对大纲 JSON 做硬性校验，返回问题列表。
    校验项：
      1. 必需顶层字段
      2. 每节课必需字段
      3. lesson_task 不能含书面作业关键词
    """
    issues = []

    # ── 1. 顶层字段 ──
    required_top = {"level", "total_lessons", "overarching_objective", "final_task", "lessons"}
    missing_top = required_top - set(outline.keys())
    if missing_top:
        issues.append(f"[结构] 大纲缺少顶层字段: {', '.join(sorted(missing_top))}")

    lessons = outline.get("lessons", [])
    if not lessons:
        issues.append("[结构] 大纲 lessons 为空")
        return issues

    # ── 2. 每节课字段 ──
    required_lesson = {"lesson_number", "lesson_name", "objective", "vocabulary", "functional_language", "lesson_task"}
    for ls in lessons:
        num = ls.get("lesson_number", "?")
        missing_ls = required_lesson - set(ls.keys())
        if missing_ls:
            issues.append(f"[结构] Lesson {num} 缺少字段: {', '.join(sorted(missing_ls))}")

    # ── 3. 书面作业红线 ──
    written_keywords = ["write an email", "draft an outline", "create slides", "write a report",
                        "write a letter", "write a paragraph", "write a summary"]
    for ls in lessons:
        task = (ls.get("lesson_task") or "").lower()
        for kw in written_keywords:
            if kw in task:
                issues.append(f"[红线] Lesson {ls.get('lesson_number', '?')} 的 lesson_task 包含书面作业: '{kw}'")

    return issues


# ── 工具函数 ──────────────────────────────────────────────

def find_unit_dir(path_arg: str | None) -> Path:
    """
    确定要测试的 Unit 文件夹：
    - 传入路径则直接用
    - 否则自动选 output/ 下最新修改的 Unit_* 文件夹
    """
    if path_arg:
        p = Path(path_arg)
        if not p.is_dir():
            raise FileNotFoundError(f"Not a directory: {p}")
        return p

    output_dir = Path("output")
    candidates = sorted(
        [d for d in output_dir.iterdir() if d.is_dir() and d.name.startswith("Unit_")],
        key=lambda d: d.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No Unit_* folders found in output/")
    return candidates[0]


def pick_random_lessons(unit_dir: Path, n: int = 3) -> list[Path]:
    """随机选取 n 份 lesson JSON（排除 unit_outline.json）"""
    lessons = [f for f in unit_dir.glob("L*.json")]
    if len(lessons) <= n:
        return lessons
    return random.sample(lessons, n)


def call_ai(prompt: str, max_retries: int = 3) -> str:
    import time
    for attempt in range(max_retries):
        try:
            response = client.chat.completions.create(
                model=AI_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
                max_tokens=4096,
            )
            return response.choices[0].message.content.strip()
        except Exception as e:
            if "429" in str(e) and attempt < max_retries - 1:
                wait = 30 * (attempt + 1)
                print(f"  [RATE LIMIT] 429 错误，等待 {wait}s 后重试 ({attempt+1}/{max_retries})...")
                time.sleep(wait)
            else:
                raise


def save_result(unit_dir: Path, test_name: str, content: str):
    qa_dir = unit_dir / "_qa"
    qa_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%m%d_%H%M%S")
    out_path = qa_dir / f"{test_name}_{timestamp}.txt"
    out_path.write_text(content, encoding="utf-8")
    print(f"  [OK] Saved: {out_path}")
    return out_path


# ── 测试1：大纲质量评估 ────────────────────────────────────

def test_outline(unit_dir: Path) -> str | None:
    print("\n" + "="*60)
    print("  TEST 1: Unit Outline QA")
    print("="*60)

    outline_path = unit_dir / "unit_outline.json"
    if not outline_path.exists():
        print(f"  [ERROR] unit_outline.json not found in {unit_dir}")
        return None

    outline = json.loads(outline_path.read_text(encoding="utf-8"))

    # ── 程序化校验（先于 AI）──
    prog_issues = programmatic_check_outline(outline)
    if prog_issues:
        print("\n  ⚙️  程序化校验发现问题:")
        for iss in prog_issues:
            print(f"    ❌ {iss}")
    else:
        print("\n  ⚙️  程序化校验: ✅ 全部通过")

    outline_text = json.dumps(outline, ensure_ascii=False, indent=2)
    prompt = OUTLINE_QA_PROMPT.replace("{{OUTLINE}}", outline_text)

    print("  [AI] Evaluating outline...")
    result = call_ai(prompt)

    # 把程序化校验结果附加到 AI 结果前面
    if prog_issues:
        prog_header = "⚙️ 程序化校验结果（硬性检查）:\n" + "\n".join(f"  ❌ {i}" for i in prog_issues) + "\n\n"
    else:
        prog_header = "⚙️ 程序化校验结果: ✅ 全部通过\n\n"
    result = prog_header + result

    print("\n" + result)
    save_result(unit_dir, "outline_qa", result)
    return result


# ── 测试2：课件内容评估 ────────────────────────────────────

def test_lessons(unit_dir: Path) -> list[str]:
    print("\n" + "="*60)
    print("  TEST 2: Lesson Content QA (3 random lessons)")
    print("="*60)

    outline_path = unit_dir / "unit_outline.json"
    outline = json.loads(outline_path.read_text(encoding="utf-8"))
    outline_lessons = {l["lesson_number"]: l for l in outline.get("lessons", [])}

    lesson_files = pick_random_lessons(unit_dir, n=3)
    if not lesson_files:
        print("  [ERROR] No lesson JSON files found")
        return []

    # ── Phase 1: Programmatic checks (fast, serial) ──
    tasks = []  # (file, lesson_num, lesson_outline, lesson_content, prog_issues, prompt)
    for f in lesson_files:
        m = re.match(r"L(\d+)", f.name)
        lesson_num = int(m.group(1)) if m else None

        lesson_outline = outline_lessons.get(lesson_num, {})
        lesson_content = json.loads(f.read_text(encoding="utf-8"))

        prog_issues = programmatic_check_lesson(lesson_content, lesson_outline)
        if prog_issues:
            print(f"\n  ⚙️  {f.name} 程序化校验发现问题:")
            for iss in prog_issues:
                print(f"    ❌ {iss}")
        else:
            print(f"\n  ⚙️  {f.name} 程序化校验: ✅ 全部通过")

        prompt = (
            LESSON_QA_PROMPT
            .replace("{{LESSON_OUTLINE}}", json.dumps(lesson_outline, ensure_ascii=False, indent=2))
            .replace("{{LESSON_CONTENT}}", json.dumps(lesson_content, ensure_ascii=False, indent=2))
        )
        tasks.append((f, lesson_num, prog_issues, prompt))

    # ── Phase 2: AI evaluation (parallel) ──
    from concurrent.futures import ThreadPoolExecutor

    def _eval_lesson(task):
        f, lesson_num, prog_issues, prompt = task
        print(f"\n  [AI] Evaluating {f.name}...")
        ai_result = call_ai(prompt)
        if prog_issues:
            prog_header = f"⚙️ {f.name} 程序化校验结果:\n" + "\n".join(f"  ❌ {i}" for i in prog_issues) + "\n\n"
        else:
            prog_header = f"⚙️ {f.name} 程序化校验结果: ✅ 全部通过\n\n"
        return f, lesson_num, prog_header + ai_result

    results = []
    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = [executor.submit(_eval_lesson, t) for t in tasks]
        for future in futures:
            f, lesson_num, result = future.result()
            print("\n" + result)
            label = f"lesson_qa_L{lesson_num}" if lesson_num else f"lesson_qa_{f.stem}"
            save_result(unit_dir, label, result)
            results.append(result)

    return results


# ── Excel 日志 ────────────────────────────────────────────

HEADERS = ["序号", "日期", "工具版本", "测试级别", "测试主题", "课包大小",
           "课包大纲JSON", "样课JSON", "QA大纲反馈", "QA样课反馈", "是否通过测试"]

LOG_PATH = Path(__file__).parent / "output" / "qa_log.xlsx"


def _get_tool_version() -> str:
    changelog = Path(__file__).parent / "CHANGELOG.md"
    if not changelog.exists():
        return "unknown"
    for line in changelog.read_text(encoding="utf-8").splitlines():
        m = re.match(r"^###\s+(v\d+\.\d+)", line)
        if m:
            return m.group(1)  # CHANGELOG is newest-first, return first match
    return "unknown"


def _is_pass(text: str) -> bool:
    return "🟢" in text and "完美通过" in text


def append_qa_log(unit_dir: Path, outline_result: str, lesson_results: list[str]):
    outline_path = unit_dir / "unit_outline.json"
    outline = json.loads(outline_path.read_text(encoding="utf-8"))

    level = outline.get("level", "")
    topic = outline.get("overarching_objective", "")[:80]
    lesson_count = len(outline.get("lessons", []))
    tool_version = _get_tool_version()

    outline_json_content = outline_path.read_text(encoding="utf-8")
    lesson_json_contents = "\n---\n".join(
        f.read_text(encoding="utf-8") for f in sorted(unit_dir.glob("L*.json"))
    )

    lesson_feedback = "\n---\n".join(lesson_results)

    outline_pass = _is_pass(outline_result) if outline_result else False
    lessons_pass = all(_is_pass(r) for r in lesson_results) if lesson_results else False
    passed = "Pass" if (outline_pass and lessons_pass) else "Fail"

    # Load or create workbook
    if LOG_PATH.exists():
        wb = openpyxl.load_workbook(LOG_PATH)
        ws = wb.active
        next_row = ws.max_row + 1
        seq = next_row - 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QA Log"
        ws.append(HEADERS)
        # Style header row
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        next_row = 2
        seq = 1

    row = [
        seq,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        tool_version,
        level,
        topic,
        lesson_count,
        outline_json_content,
        lesson_json_contents,
        outline_result or "",
        lesson_feedback,
        passed,
    ]
    ws.append(row)

    # Style data row
    row_idx = next_row
    wrap_cols = {9, 10}  # QA feedback columns (1-indexed)
    for col_idx, cell in enumerate(ws[row_idx], start=1):
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in wrap_cols))
    ws.row_dimensions[row_idx].height = 80

    # Color pass/fail cell
    pass_cell = ws.cell(row=row_idx, column=11)
    if passed == "Pass":
        pass_cell.fill = PatternFill("solid", fgColor="D1FAE5")
        pass_cell.font = Font(color="065F46", bold=True)
    else:
        pass_cell.fill = PatternFill("solid", fgColor="FEE2E2")
        pass_cell.font = Font(color="991B1B", bold=True)
    pass_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    col_widths = [6, 18, 12, 10, 40, 10, 20, 40, 60, 60, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    LOG_PATH.parent.mkdir(exist_ok=True)
    # Try saving; if file is locked (e.g. open in Excel), save to a timestamped backup
    save_path = LOG_PATH
    try:
        wb.save(save_path)
    except PermissionError:
        ts = datetime.now().strftime("%m%d_%H%M%S")
        save_path = LOG_PATH.parent / f"qa_log_{ts}.xlsx"
        wb.save(save_path)
        print(f"\n  [WARN] qa_log.xlsx is open in Excel. Saved to: {save_path.name}")
    print(f"\n  [OK] QA log updated: {save_path.resolve()}")
    print(f"  Result: {passed}")


# ── 主入口 ────────────────────────────────────────────────

if __name__ == "__main__":
    unit_path_arg = sys.argv[1] if len(sys.argv) > 1 else None

    try:
        unit_dir = find_unit_dir(unit_path_arg)
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        sys.exit(1)

    print(f"\nUnit folder: {unit_dir.resolve()}")

    outline_result = test_outline(unit_dir)
    lesson_results = test_lessons(unit_dir)

    append_qa_log(unit_dir, outline_result, lesson_results)

    print("\n" + "="*60)
    print("  QA complete. Results saved to _qa/ folder.")
    print("="*60 + "\n")
